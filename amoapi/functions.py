from django.core.mail import EmailMessage
from django.core.mail.backends.smtp import EmailBackend
import time
from .models import Leads, Paints, PaintsLeads, MailPass
from .serializers import PaintsLeadsSerializer, PaintsLeadsSerializerLink, PaintsSerializer, PaintsSerializerLink, PaintsLeadsSerializerEdit
from .serializers import PaintsLeadsEmailSerializer
import logging
from django.http import HttpResponse
from .serializers import PaintsFullSerializer, PaintsLeadsFullSerializer
from django.db.models import Q
from docxtpl import DocxTemplate

logging.basicConfig(level=logging.DEBUG, filename='/home/perespimka/monyze/log.txt', format='%(asctime)s %(levelname)s %(message)s')
NAME_SWITCH = {
    "ral": "RAL",
    "ral design": "RALD",
    "ral classic": "RALC",
    "ral effect": "RALE",
    "nsc": "NCS",
    "pantone": "PNT",
    "Глянцевая 80-100%": "80/100%",
    "Полулянцевая 55-80%": "55/80%", 
    "Полуматовая 30-55%": "30/55%",
    "Матовая 15-30%": "15/30%",
    "Суперматовая 1-15%": "1-15%", 
    "Шагрень": "шгр",
    "Гладкая": "гл",
    "Антик": "ант",
    "Муар": "мр",
    "Молотковая": "мтк",
    "Шелк": "шлк",
    "Не определено": "н/о",
    }

def combine_paint_data(pl, PlSerializator, PSerializator):
    '''Возвращает словарь объединенный paints и paints_leads'''
    serializer_pl = PlSerializator(pl).data
    serializer_p = PSerializator(pl.paint).data
    serializer_p.update(serializer_pl)
    return serializer_p

def tag_creating(pl):
    '''Создание тегов'''
    res_string = f'{pl.paint.product} {pl.paint.catalog}{pl.paint.code}'
    if pl.product_type == 'sample':
        res_string = f'Образец {res_string}' 
    return res_string

def get_lead_data(req):
    '''
    Возвращаем данные по сделке
    '''
    try:
        lead = Leads.objects.get(amo_lead_id=req['lead_id'])
    except Leads.DoesNotExist:
        lead = Leads(amo_lead_id=req['lead_id'])
        lead.save()
    paints_leads = lead.paintsleads_set.all()
    sum_ = 0
    serializer = []
    if len(paints_leads) > 0:
        for pl in paints_leads:
            try:
                sum_ += pl.price * pl.vol # Тут уточнить по обязательности заполнения полей
            except:
                pass
            combined = combine_paint_data(pl, PaintsLeadsSerializer, PaintsSerializer)
            if pl.new_lead:
                combined['new_lead_id'] = pl.new_lead.amo_lead_id
            else:
                combined['new_lead_id'] = None
            combined['tag'] = tag_creating(pl)
            serializer.append(combined)
        
    else:
        serializer = 'none' # На фронте так надо. Там ждут именно строку 'none', а не json'овский null
    result = {
        'id': lead.id,
        'amo_lead_id': lead.amo_lead_id,
        'sum': sum_,
        'lead_paints_info': serializer
    }
    return result

def name_switch(name):
    if name in NAME_SWITCH:
        return NAME_SWITCH[name]
    return name

def get_paint_name(req):
    return (name_switch(req['catalog']) + req['code'] + '-' + req['basis'] + '-' + name_switch(req['shine']) +
                    '-' + name_switch(req['facture'])
    )

def attach_goods(req):
    '''
    Запись краски и дополнительных данных в paints и paints_leads
    '''
    if all((req['product'], req['basis'], req['catalog'], req['code'], req['shine'], req['facture'])):
        serialize_p = PaintsSerializerLink(data=req)
        logging.debug(req)
        if serialize_p.is_valid():
            name = get_paint_name(req)
            try:
                paint = Paints.objects.get(name=name, product=serialize_p.validated_data['product'], basis=serialize_p.validated_data['basis'],
                                           catalog=serialize_p.validated_data['catalog'], code=serialize_p.validated_data['code'], shine=serialize_p.validated_data['shine'],
                                           facture=serialize_p.validated_data['facture']
                )
            except:
                paint = serialize_p.save(name=name)
        else:
            logging.debug(serialize_p.errors)
        try:
            lead = Leads.objects.get(amo_lead_id=req['lead_id'])
        except:
            lead = Leads(amo_lead_id=req['lead_id']) 
            lead.save()
        serialize_pl = PaintsLeadsSerializerLink(data=req)
        if serialize_pl.is_valid():
            serialize_pl.save(lead=lead, paint=paint)
        else:
            logging.debug(serialize_pl.errors)
        
        pls = lead.paintsleads_set.all() # Все записи в paints_leads у текущего лида
        result = []
        for pl in pls:
            res_string = f'{pl.paint.product} {pl.paint.catalog}{pl.paint.code}'
            if pl.product_type == 'sample':
                res_string = f'Образец {res_string}'            
            result.append(res_string)      
        result = ','.join(result)

        return {'status': 'done', 'tags': result} # Так просит фронт 

def del_paintsleads_rec(req):
    '''
    Удаляем запись в paints_leads
    '''
    try:
        pl = PaintsLeads.objects.get(id=req['paints_leads_id'])
    except:
        return {'status': 'not ok'}
    pl.delete()
    lead = Leads.objects.get(amo_lead_id=req['lead_id'])
    result = []
    for pl in lead.paintsleads_set.all():
        res_string = f'{pl.paint.product} {pl.paint.catalog}{pl.paint.code}'
        result.append(res_string)
    result = ','.join(result)
    return {'status': 'done', 'tags': result}
    
def edit_paint(req):
    '''
    Редактирование краски
    '''
    if all((req['lead_id'], req['paints_leads_id'])):
        name = get_paint_name(req)
        serializer_p = PaintsSerializerLink(data=req)
        logging.debug(req)
        if serializer_p.is_valid():
            try:
                paint = Paints.objects.get(name=name, product=req['product'], basis=req['basis'],
                                           catalog=req['catalog'], code=req['code'], 
                                           shine=req['shine'], facture=req['facture']
                                           
                )
            except:
                paint = serializer_p.save(name=name)
        else:
            logging.debug(serializer_p.errors)
        lead = Leads.objects.get(amo_lead_id=req['lead_id'])
        try:
            pl = PaintsLeads.objects.get(id=req['paints_leads_id']) #!!!
        except:
            return {'status': 'omg  paints_leads id not found'}
        serializer_pl = PaintsLeadsSerializerEdit(pl, data=req)
        if serializer_pl.is_valid():
            serializer_pl.save(paint=paint)
        else:
            logging.debug(serializer_pl.errors)
        result = []
        for pl in lead.paintsleads_set.all():
            res_string = f'{pl.paint.product} {pl.paint.catalog}{pl.paint.code}'
            if pl.product_type == 'sample':
                res_string = f'Образец {res_string}'
            result.append(res_string)
        result = ','.join(result)
        return {'status': 'done', 'tags': result}

def get_paint_info(req):
    '''
    Данные по краске. Запрос на /paints
    '''
    if 'paints_leads_id' in req:
        try:
            pl = PaintsLeads.objects.get(id=req['paints_leads_id'])
        except:
            return {'status': 'paints_leads_id not found'}
        return combine_paint_data(pl, PaintsLeadsFullSerializer, PaintsFullSerializer)
    elif 'paint_id' in req:
        try:
            paint = Paints.objects.get(id=req['paint_id'])
        except:
            return {'status': 'paint_id not found'}
        return PaintsFullSerializer(paint).data

def paint_search(req):
    '''
    ПОиск краски. 
    '''
    query = req['query']
    q = (
        Q(name__icontains=query) | Q(product__icontains=query) | Q(basis__icontains=query) | Q(catalog__icontains=query) |
        Q(code__icontains=query) | Q(shine__icontains=query) | Q(facture__icontains=query)
    )
    result = []
    for paint in Paints.objects.filter(q):
        result.append({'id': paint.id, 'name': paint.name})
    return result
#*******   

def data_to_xlsx(req, paints_to_xlsx):
    '''
    Формируем xlsx файлы по шаблону. Одна краска - один файл. Возвращаем список имен файлов
    '''
    from openpyxl import load_workbook
    data_manager = req['data_manager']
    count = 0
    fnames = []
    if req['state'] == 'send_lab':
        for paint in paints_to_xlsx:
            wb = load_workbook('sample_template.xlsx')
            ws = wb.active
            if paint['paid_sample'] == 1:
                ws['B7'] = 'Да'
            else:
                ws['B7'] = 'Нет'
            ws['B6'] = data_manager.get('man_comp_payer') # ИНН
            ws['B5'] = data_manager.get('man_comp_name') # Имя компании
            ws['B3'] = data_manager.get('man_name') # Имя менеджера
            ws['B4'] = time.strftime('%d-%m-%Y', time.localtime()) # Дата оформления заявки
            ws['B9'] = paint['name']
            ws['B10'] = paint['product']
            ws['B11'] = paint['basis']
            ws['B12'] = paint['catalog']
            ws['B13'] = paint['code']
            ws['B14'] = paint['shine']
            ws['B15'] = paint['facture']
            ws['B16'] = paint.get('temperature')
            ws['B17'] = paint.get('applying')
            ws['B19'] = paint.get('surface_type')
            ws['B20'] = paint.get('surface_thin')
            ws['B22'] = str(paint.get('panels'))
            ws['B23'] = paint.get('powder')
            ws['B24'] = paint.get('client_sample')
            ws['B25'] = paint.get('comment')
            ws['D10'] = paint.get('postforming')
            ws['D11'] = paint.get('metallic')
            ws['D12'] = paint.get('chameleon')   
            ws['D13'] = paint.get('antibacterial')
            ws['D14'] = paint.get('antigraffiti')
            ws['D15'] = paint.get('architect')
            ws['D16'] = paint.get('zinc')
            ws['D18'] = paint.get('sublim')
            fname = f'{paint["name"]}_{count}.xlsx'.replace('/', '_')
            fname = 'mail_files/' + fname
            count += 1
            fnames.append(fname)
            wb.save(fname)
    elif req['state'] == 'set_score':
        for paint in paints_to_xlsx:
            wb = load_workbook('invoice_template.xlsx')
            ws = wb.active
            ws['B7'] = paint.get('vol')
            ws['B8'] = paint.get('price')
            ws['B9'] = paint.get('delivery_terms')
            ws['B10'] = paint.get('delivery_date')
            ws['B6'] = data_manager.get('man_comp_payer') # ИНН
            ws['B5'] = data_manager.get('man_comp_name') # Имя компании
            ws['B3'] = data_manager.get('man_name') # Имя менеджера
            ws['B4'] = time.strftime('%d-%m-%Y', time.localtime()) # Дата оформления заявки
            ws['B12'] = paint['name']
            ws['B13'] = paint['product']
            ws['B14'] = paint['basis']
            ws['B15'] = paint['catalog']
            ws['B16'] = paint['code']
            ws['B17'] = paint['shine']
            ws['B18'] = paint['facture']
            ws['B19'] = paint.get('temperature')
            ws['B20'] = paint.get('applying')
            ws['B22'] = paint.get('surface_type')
            ws['B23'] = paint.get('surface_thin')
            ws['B25'] = paint.get('comment')
            ws['D13'] = paint.get('postforming')
            ws['D14'] = paint.get('metallic')
            ws['D15'] = paint.get('chameleon')   
            ws['D16'] = paint.get('antibacterial')
            ws['D17'] = paint.get('antigraffiti')
            ws['D18'] = paint.get('architect')
            ws['D19'] = paint.get('zinc')
            ws['D21'] = paint.get('sublim')
            fname = f'{paint["name"]}_{count}.xlsx'.replace('/', '_')
            fname = 'mail_files/' + fname
            count += 1
            fnames.append(fname)
            wb.save(fname)        
    return fnames
    


def send_mail_to(req, emails, attachments, subject, body): # Переписать attach как список имен файлов
    '''Отправка письма с кодировкой'''
    import base64

    #sample_data = data['sample_data'].copy()
    manager_mail = req['manager_mail']
    if manager_mail == 'stardustpaintsamo@mail.ru':
        manager_mail = 'soloviev@stardustpaints.ru'
    try:
        user = MailPass.objects.get(email=manager_mail)
    except:
        logging.debug('Не удалось найти данные почтового адреса менеджера для отправки письма')
        return None
    
    manager_mail_pass = base64.b64decode(user.password.encode()).decode('utf-8')
    em = EmailMessage(subject=subject, body=body,
                      to=emails, from_email=manager_mail
    )
    em.content_subtype = 'html'
    for attach in attachments:
        em.attach_file(attach) 
    EmailBackend(
        host = 'mail.stardustpaints.ru',
        username = manager_mail,
        password = manager_mail_pass,
        #use_ssl = True,
        use_tls = True,
        port = 587
    ).send_messages([em])    

def send_mail_to_lab_prod(req):
    '''
    Отправка письма в лабораторию, офис
    '''
    if req['state'] == 'send_lab':
        emails = ['s.dmitrievlol@yandex.ru', 'soloviev357@gmail.com', 'ts@stardustpaints.ru'] # Изменить на лабу
        status_value = 1
        subject = 'Создать образец'
        msg = 'Просьба создать образец, данные во вложении'
    elif req['state'] == 'send_prod':
        emails = ['s.dmitrievlol@yandex.ru', 'soloviev357@gmail.com', 'ts@stardustpaints.ru'] # Изменить на бэкофис
        status_value = 2
        subject = 'В производство'
        msg = 'Отправить в производство, данные во вложении'
    elif req['state'] == 'set_score':
        emails = ['s.dmitrievlol@yandex.ru', 'soloviev357@gmail.com', 'ts@stardustpaints.ru'] # Изменить на бэкофис
        status_value = 2
        subject = 'Сделать счет'
        msg = 'Создать счет, данные во вложении'        

    #new_lead = Leads(amo_lead_id=req['new_id'])
    try:
        new_lead = Leads.objects.get(amo_lead_id=req['new_id'])
    except:
        new_lead = None
    paints_to_xlsx = []
    #Если сделка с таким id уже существует, тогда просто обновляем статус у красок
    if new_lead: 
        for pl_id in req['id']:
            pl = PaintsLeads.objects.get(id=pl_id)
            pl.status = status_value
            pl.save()
            paints_to_xlsx.append(combine_paint_data(pl, PaintsLeadsEmailSerializer, PaintsFullSerializer))
    #Если сделки нет, создаем новую сделку и копируем в нее краски. В старых красках в поле new_lead пишем новый айди сделки
    else:
        new_lead = Leads(amo_lead_id=req['new_id'])
  
        try:
            new_lead.save()
        except:
            logging.debug(f'Новый лид уже существует ({new_lead.amo_lead_id} , state: {req["state"]})')

        
        for pl_id in req['id']:
            pl = PaintsLeads.objects.get(id=pl_id)
            pl.status = status_value
            pl.new_lead = new_lead
            pl.save()
            pl.id = None
            pl.lead = new_lead
            pl.new_lead = None
            pl.save()
            paints_to_xlsx.append(combine_paint_data(pl, PaintsLeadsEmailSerializer, PaintsFullSerializer))
    # Ниже меняем функцию по обработке, которая будет возвращать список имен файлов

    attachments = data_to_xlsx(req, paints_to_xlsx)
    send_mail_to(req, emails, attachments, subject, msg)

def email_from_contacts(req):
    '''
    Данные из запроса возвращаем в виде кортежа (мейл, имя, должность)
    '''
    contacts = req['data_manager']['man_comp_users']
    email = req['data_manager']['client_mail']
    name = req['data_manager']['client_name']
    return email, name

def send_cp(req):
    '''Отправка коммерческого предложения'''
    from .html_sign import sign
    context = {}
    contacts = email_from_contacts(req)
    logging.debug(f'contacts tuple: {contacts}')
    if contacts:
        email, context['fio_from_card'] = contacts
        context['company_name'] = req['data_manager']['man_comp_name']
        context['manager_name'] = req['data_manager']['man_name'] + ' ' + req['data_manager']['man_last_name']
        context['manager_email'] = req['data_manager']['man_login']
        context['manager_phone'] = req['data_manager']['man_phone_number']
        context['tbl'] = []
        for pl_id in req['id']:
            pl_table = {}
            pl = PaintsLeads.objects.get(id=int(pl_id))
            applying = pl.applying
            if applying and applying != 'Не определено':
                applying = f', применение {pl.applying}'
            params = (pl.paint.catalog, pl.paint.code, pl.paint.basis, pl.paint.facture, pl.paint.shine, applying)
            params = [param for param in params if param != 'Не определено']
            pl_table['name'] = ' '.join(params)
            '''
            pl_table['name'] = (f'{pl.paint.catalog} {pl.paint.code} {pl.paint.basis} {pl.paint.facture} '
                                f'{pl.paint.shine}, применение {pl.applying}'
            )
            '''
            pl_table['vol'] = str(pl.vol)
            pl_table['price'] = str(pl.price)
            context['tbl'].append(pl_table)
        doc = DocxTemplate('/home/perespimka/monyze/cp_stardustpaints.docx')
        doc.render(context)
        now = time.strftime('%d-%m-%Y', time.localtime())
        fname = f'mail_files/cp_to_{context["company_name"].lower()}_{now}.docx'
        doc.save(fname)
        body = sign.format(**context)
        send_mail_to(req, [email, 's.dmitrievlol@yandex.ru', 'soloviev357@gmail.com', 'dmitrievs@stardustpaints.ru'], [fname],
                     'Коммерческое предложение Stardustpaints', body   
        )


if __name__ == "__main__":
    print(name_switch('RAL'))